from __future__ import annotations

import csv
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook


def _cell_to_str(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return s


def csv_to_table_text(path: str, max_rows: int = 80) -> str:
    p = Path(path)
    rows = []
    with p.open("r", encoding="utf-8", errors="ignore", newline="") as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            if i >= max_rows:
                break
            rows.append([_cell_to_str(x) for x in row])

    # salida estable: TAB separado
    return "\n".join(["\t".join(r) for r in rows])


def xlsx_to_table_text(path: str, max_rows: int = 80, sheet: int = 0) -> str:
    p = Path(path)
    wb = load_workbook(filename=str(p), read_only=True, data_only=True)
    ws = wb.worksheets[sheet]

    lines = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        cells = [_cell_to_str(v) for v in row]
        # recorta trailing vacíos para estabilidad
        while cells and cells[-1] == "":
            cells.pop()
        lines.append("\t".join(cells))

    wb.close()
    return "\n".join(lines)
